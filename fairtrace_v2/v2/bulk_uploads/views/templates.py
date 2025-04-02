import csv
import json
from plistlib import InvalidFileException
from typing import get_args

import openpyxl
from common import library
from common.drf_custom.mixins import inject_node
from common.drf_custom.views import IdencodeObjectViewSetMixin
from common.library import camel_to_underscore, decode, success_response
from django.http import HttpResponse
from openpyxl.reader.excel import load_workbook
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from v2.accounts import permissions as user_permissions
from v2.supply_chains import permissions as sc_permissions

from ...supply_chains.constants import NODE_TYPE_FARM
from ...supply_chains.models import SupplyChain
from ..constants import TEMPLATE_TYPE_CONNECTION, TEMPLATE_TYPE_TXN
from ..models import DataSheetTemplate
from ..schemas.farmer_upload_schema import FarmerUploadSchema
from ..schemas.transaction_upload_schema import TransactionUploadSchema
from ..serializers.templates import DataSheetTemplateSerializer


class DataSheetTemplateViewSet(
    IdencodeObjectViewSetMixin, viewsets.ModelViewSet
):
    """Data sheet template view set.

    This ViewSet handles operations related to data sheet templates. It
    provides CRUD functionality for DataSheetTemplate instances.
    Additionally, it supports filtering and soft deletion of instances.
    """

    queryset = DataSheetTemplate.objects.all()
    serializer_class = DataSheetTemplateSerializer

    permission_classes = (
        user_permissions.IsAuthenticated,
        sc_permissions.HasNodeAccess,
    )

    def get_queryset(self):
        """Get the filtered queryset.

        This method filters the queryset based on the query parameters provided
        in the request.

        Returns:
            QuerySet: The filtered queryset.
        """
        return super().get_queryset().filter_by_query_params(
            self.request).order_by("-is_system_template")

    def create(self, request, *args, **kwargs):
        """Create a new DataSheetUpload instance.

        This method creates a new DataSheetUpload instance and injects
        the current node as the source. It then delegates the handling
        of the request to the parent function.
        """
        file = request.data.get("file")
        header_row = self.find_header_row(file) if file else None
        inject_node(request, **kwargs)
        res = super().create(request, *args, **kwargs)
        res.data["header_row"] = header_row
        res.data["data_row"] = header_row + 1 if header_row else None
        return res

    def update(self, request, *args, **kwargs):
        """Update the DataSheetTemplate instance.

        This method updates the DataSheetTemplate instance and injects
        the current node as the source. It then delegates the handling
        of the request to the parent function.
        """
        inject_node(request, **kwargs)
        return super().update(request, *args, **kwargs)

    def perform_destroy(self, instance):
        """Soft delete the DataSheetTemplate instance.

        This method performs a soft delete of the DataSheetTemplate
        instance by setting the 'is_deleted' field to True and saving
        the instance.
        """
        instance.is_deleted = True
        instance.save()

    @action(methods=["get"], detail=True, url_path="preview")
    def preview(self, request, *args, **kwargs):
        """
        Preview the content of the uploaded file.

        This method is an action endpoint for previewing the content of the
        uploaded file associated with the detail record. It handles HTTP GET
        requests sent to the endpoint "preview".
        """

        # Get the instance associated with the detail record
        instance = self.get_object()

        # Read the content of the file
        file = instance.file._get_file()  # noqa

        # Determine the file format and read the data accordingly
        if instance.file.name.endswith(".csv"):
            data = self._read_csv_file(file)
        elif instance.file.name.endswith(".xlsx"):
            data = self._read_excel_file(file)
        elif instance.file.name.endswith(".json"):
            data = json.load(file)
        else:
            raise ValidationError(
                "Got wrong file format. Check file extension.")

        # Return the preview data in the response
        return success_response(data)

    @action(methods=["get"], detail=True, url_path="sample-file")
    def sample_file(self, request, *args, **kwargs):
        """Get a sample file.

        This method returns a sample file for the DataSheetTemplate instance.

        Returns:
            FileResponse: The sample file.
        """
        instance = self.get_object()
        file = instance.file._get_file()  # noqa

        supply_chain = request.query_params.get("supply_chain")

        if supply_chain:
            supply_chain = SupplyChain.objects.get(id=decode(supply_chain))

        # Create a workbook object
        wb = openpyxl.load_workbook(file)

        # Select the desired sheet (you can choose by name or index)
        sheet = wb.active

        # Clear all rows except the header row
        for row in list(sheet.iter_rows())[instance.data_row:]:
            for cell in row:
                cell.value = None

        if instance.is_system_template:
            # prefill farmer data if trace template
            self._prefill_farmer_data(sheet, instance, supply_chain)

        # Save the new workbook to a BytesIO object
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Set the appropriate response headers
        response = HttpResponse(
            content_type=('application/vnd.openxmlformats-officedocument'
                          '.spreadsheetml.sheet'))
        response['Content-Disposition'] = 'attachment; filename=output.xlsx'

        # Write the BytesIO object to the response
        response.write(output.getvalue())
        return response

    @action(methods=["get"], detail=False, url_path="schema-fields")
    def get_schema_fields(self, request, *args, **kwargs):
        """Get schema fields.

        Retrieves the schema fields from the `FarmerUploadSchema` class and
        returns the field names, field types, and whether they are required or
        not.

        Returns:
            A sorted JSON response containing the schema field data.
        """
        template_type = request.query_params.get("type")

        if int(template_type) == TEMPLATE_TYPE_TXN:
            schema = TransactionUploadSchema
        elif int(template_type) == TEMPLATE_TYPE_CONNECTION:
            schema = FarmerUploadSchema
        else:
            raise ValidationError("Template type not provided or not "
                                  "supported")
        mandatory_fields = schema.get_mandatory_fields().keys()
        data = []

        def get_field_type(field_arg):
            # get the type of the schema field
            first_arg = get_args(field_arg)[0]
            if not get_args(first_arg):
                return first_arg
            return get_field_type(first_arg)

        for field, field_type in schema.get_fields().items():
            _type = get_field_type(field_type)
            field_data = {
                "field": field,
                "label": field.replace("_", " ").title(),
                "field_type": camel_to_underscore(_type.__name__),
                "required": False,
            }

            if field in mandatory_fields:
                field_data["required"] = True
            data.append(field_data)
        sorted_data = sorted(data, key=lambda x: x["required"], reverse=True)

        return success_response(data=sorted_data)

    @staticmethod
    def find_header_row(file):
        """
        Find the header row in an Excel file.

        This function takes an Excel file and searches for the row where the
        header information is located. The header row is determined by finding
        columns with at least one non-empty cell and then searching for the
        first row with all non-empty cells in those columns.

        Parameters:
        file (str): The path to the Excel file.

        Returns:
        int or None: The row number where the header is found, or None if the
        header is not found.
        """
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # Find the dimensions of the sheet (number of rows and columns)
        max_row = sheet.max_row
        max_column = sheet.max_column

        # Find columns with at least one non-empty cell
        non_empty_columns = []
        for col in range(1, max_column + 1):
            column_data = [sheet.cell(row=row, column=col).value for row in
                           range(1, max_row + 1)]
            if any(cell_value is not None for cell_value in column_data):
                non_empty_columns.append(col)

        # If no non-empty columns are found, return None
        if not non_empty_columns:
            return None

        def is_empty_row(row, columns):
            """
            Find the first row with all non-empty cells in the non-empty
            columns.
            """
            if not columns:
                return None
            for row in range(1, max_row + 1):
                if all(sheet.cell(row=row, column=_col).value is not None
                       for _col in columns):
                    return row
            return is_empty_row(
                row, columns[1:])  # Recursively call is_empty_row

        return is_empty_row(1, non_empty_columns)

    @staticmethod
    def _read_excel_file(file):
        """
        Read data from an Excel file and return as a 2D array.

        This function reads the content of an Excel file and returns it as a
        2D array (list of lists). Each row in the Excel sheet is represented as
        a list, and the entire data is represented as a list of rows.

        Parameters:
        file (BinaryIO): The binary file object of the Excel file to be read.

        Returns:
        List[List]: A 2D array representing the data from the Excel file.
                    Each inner list represents a row in the Excel sheet.
        """

        # Load the Excel workbook
        try:
            workbook = load_workbook(file)
        except InvalidFileException:
            raise InvalidFileException(
                "The provided file is not a valid Excel file.")

        # Get the active sheet from the workbook
        sheet = workbook.active

        # Read data from the sheet and store it in a 2D array
        data_2d_array = []
        for row in sheet.iter_rows(values_only=True):
            data_2d_array.append(list(row))

        last_row_index = None
        # Find last row with at least one non-null value
        for index, row in enumerate(data_2d_array):
            if any(cell is not None for cell in row):
                last_row_index = index

        last_column_index = None
        # Find last column with at least one non-null value
        for i in range(len(data_2d_array[0])):
            if any(row[i] is not None for row in data_2d_array):
                last_column_index = i

        return [
            row[:last_column_index + 1]
            for row in data_2d_array[:10]]

    @staticmethod
    def _read_csv_file(file):
        """
        Read data from a CSV file and return as a 2D array.

        This function reads the content of a CSV file and returns it as a
        2D array (list of lists). Each row in the CSV file is represented as a
        list, and the entire data is represented as a list of rows.

        Parameters:
        file (TextIO): The text file object of the CSV file to be read.

        Returns:
        List[List]: A 2D array representing the data from the CSV file. Each
                    inner list represents a row in the CSV file.
        """

        # Initialize an empty list to store the data
        data_2d_array = []

        try:
            # Use the csv.reader to read data from the CSV file
            reader = csv.reader(file)

            # Iterate through each row in the CSV file and append it to the
            # data list
            for row in reader:
                data_2d_array.append(row)

        except csv.Error as e:
            raise csv.Error(
                f"Error occurred while reading the CSV file: {str(e)}")

        return data_2d_array

    def _prefill_farmer_data(self, sheet, instance, supply_chain=None):
        """
        Prefill farmer data into the Excel sheet.

        This method takes an Excel sheet and an instance of the
        DataSheetTemplate model and pre-fills the farmer data into the sheet
        based on the node. The data is fetched from the suppliers of the
        specified node's supply chain and is populated into the specified
        rows and columns.

        Parameters:
        sheet (openpyxl.worksheet.Worksheet): The Excel sheet to prefill.
        instance (DataSheetUpload): An instance of the DataSheetUpload model.

        Returns:
        None
        """
        node = self.kwargs.get("node")
        if not node:
            return

        # Get farmers from the specified node's supply chain
        farmers = node.get_suppliers(
            supply_chain=supply_chain).filter(type=NODE_TYPE_FARM)

        # Get indexed columns from the template
        indexed_fields = instance.indexed_fields

        # Iterate through rows and update data based on the queryset
        for row_number, row_data in enumerate(
                farmers, start=instance.data_row + 1):
            for column_index, field in indexed_fields.items():
                value = getattr(row_data.farmer, field, None)
                if column_index == 0:
                    value = row_data.idencode
                if value is None and field == "connection_type":
                    value = self._get_primary_operator(row_data, supply_chain)
                if field == "phone":
                    _, value = library.split_phone(value)
                if field == "country_code":
                    value, _ = library.split_phone(row_data.farmer.phone)

                # Populate the cell with the corresponding value
                sheet.cell(
                    row=row_number, column=column_index + 1).value = value

    @staticmethod
    def _get_primary_operator(instance, supply_chain):
        """
        Get the name of the primary operator associated with the instance
        within the specified supply chain.

        This function retrieves the name of the primary operator for the
        given instance within the context of the provided supply chain. If a
        supply chain is specified, it attempts to find the primary operator
        associated with the supply chain through a NodeSupplyChain
        relationship. If found, the name of that primary operator is
        returned. If no supply chain or no matching primary operator is
        found, the name of the instance's primary operator (if available) is
        returned. If neither supply chain nor primary operator is found,
        an empty string is returned.

        Args:
            instance: An instance for which to retrieve the primary
            operator. supply_chain: The supply chain within which to search for
            the primary operator (can be None).

        Returns:
            str: The name of the primary operator associated with the
            instance in the specified supply chain, or the name of the
            instance's primary operator if found. Returns an empty string if no
            primary operator is found.
        """
        if supply_chain:
            nsc = instance.nodesupplychain_set.filter(
                supply_chain=supply_chain,
                primary_operation__isnull=False,
            ).first()
            if nsc:
                return nsc.primary_operation.name
        return (instance.primary_operation.name
                if instance.primary_operation
                else "")
