import copy

from common import library as comm_lib
from common.exceptions import BadRequest
from openpyxl import load_workbook
from openpyxl import utils as xl_utils
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
from v2.supply_chains import constants
from v2.supply_chains.models import Operation

from .cells import DependantChoiceCell
from .constants import VALUE_CHANGED
from .constants import VALUE_NEW
from .constants import VALUE_UNCHANGED


class ExcelRow:
    """Base Class for defining the structure and types of cells in a row."""

    initial_data: dict = {}
    index = 1
    model_object = None

    class Meta:
        name = ""
        fields: list = []
        related_model = None
        id_field = "id"

    def __init__(
        self, row_data=None, initial_data=None, index=1, *args, **kwargs
    ):
        data = {}
        if row_data:
            for c in row_data:
                data[c.column_letter] = c.value
            if self.Meta.related_model:
                id_field = getattr(self, self.Meta.id_field)
                object_id = comm_lib._decode(data[id_field.column])
                if object_id:
                    self.model_object = self.Meta.related_model.objects.get(
                        id=object_id
                    )
            for field_name in self.Meta.fields:
                setattr(
                    self, field_name, copy.deepcopy(getattr(self, field_name))
                )
                field = getattr(self, field_name)
                field_value = data[field.column] if data[field.column] else ""
                field.set_value(field_value)
                field.model_object = self.model_object
                if self.model_object and field.update:
                    db_field_name = (
                        field.source if field.source else field_name
                    )
                    database_value = getattr(self.model_object, db_field_name)
                    field.database_value = (
                        database_value if database_value else ""
                    )
        elif initial_data and index:
            self.index = index
            self.initial_data = initial_data

    def __str__(self):
        """To perform function __str__."""
        return ", ".join(
            [str(getattr(self, i).value) for i in self.Meta.fields]
        )

    def validate(self):
        """To perform function validate."""
        valid = True
        message = ""
        data = {}
        statuses = []
        total_issues = 0
        trans_issues = 0
        farmer_issues = 0
        for field_name in self.Meta.fields:
            field = getattr(self, field_name)
            if type(field) == DependantChoiceCell:
                selection = getattr(self, field.selection)
                check = field.validate(selection=selection)
            else:
                check = field.validate()
            valid &= check["valid"]
            message += check["message"]
            data[field_name] = check
            statuses.append(check["status"])
            total_issues += data[field_name]["issues"]["total"]
            trans_issues += data[field_name]["issues"]["transaction_issues"]
            farmer_issues += data[field_name]["issues"]["farmer_issues"]
            data["issues"] = {
                "total": total_issues,
                "transaction_issues": trans_issues,
                "farmer_issues": farmer_issues,
            }

        if VALUE_CHANGED in statuses:
            data["status"] = VALUE_CHANGED
        elif set(statuses) == {VALUE_NEW}:
            data["status"] = VALUE_NEW
        else:
            data["status"] = VALUE_UNCHANGED

        data["valid"] = valid
        data["message"] = message
        return data

    def set_value(self, sheet):
        """To perform function set_value."""
        for field_name in self.Meta.fields:
            field = getattr(self, field_name)
            if field.write:
                cell = "%s%s" % (field.column, self.index)
                sheet[cell] = self.initial_data.get(field_name, "")
        return sheet


class Excel:
    """Class to handle Excel and functions."""

    rows: list = []
    workbook = None
    file_path = ""
    data: dict = {}
    visible_columns: list = []
    data_col_start = 2000
    model_name = None

    class Meta:
        first_row = 1
        row_class = None
        backend_sheet = ""
        data_sheet = ""
        fields: list = []
        mandatory_column = None

    def define_name(self, name, options):
        """To perform function ine_name."""
        sheet = self.workbook[self.Meta.backend_sheet]
        col = xl_utils.get_column_letter(self.data_col_start)
        options_count = len(options)
        for i in range(options_count):
            cell = f"{col}{i + 1}"
            sheet[cell] = options[i]

        formula = f"${col}${1}:${col}${options_count}"
        self.workbook.create_named_range(name, sheet, formula)
        self.data_col_start += 1
        return name

    def __init__(
        self,
        workbook=None,
        file_path=None,
        data=None,
        visible_columns=None,
        *args,
        **kwargs,
    ):
        if data:
            # if data are from farmer model then save model_name as
            # farmer
            if "farmers" in data.keys():
                self.model_name = dict(constants.NODE_TYPE_CHOICES)[
                    constants.NODE_TYPE_FARM
                ]
        self.rows = []
        if visible_columns:
            self.visible_columns = visible_columns
        if workbook:
            self.workbook = workbook
            try:
                sheet = self.workbook[self.Meta.data_sheet]
            except KeyError:
                raise BadRequest("Invalid Excel file")
            col_index = (
                xl_utils.column_index_from_string(self.Meta.mandatory_column)
                - 1
            )
            for row in sheet.iter_rows(min_row=self.Meta.first_row):
                if row[col_index].value:
                    row_object = self.Meta.row_class(
                        row_data=row, visible_columns=visible_columns
                    )
                    self.rows += [row_object]
                    del row_object
                else:
                    pass
            for field_name in self.Meta.fields:
                setattr(
                    self, field_name, copy.deepcopy(getattr(self, field_name))
                )
                field = getattr(self, field_name)
                field.set_value(sheet[field.cell].value)
        elif file_path:
            self.file_path = file_path
            if data:
                self.data = data
        else:
            raise AttributeError(
                "Either 'workbook' or 'file_path' should be provided when"
                " implementing Excel."
            )

    def validate(self):
        """To perform function validate."""
        valid = True
        row_data = []
        message = ""
        statuses = []
        for row in self.rows:
            check = row.validate()
            valid &= check["valid"]
            message += check["message"]
            row_data.append(check)
        excel_data = {}
        for field_name in self.Meta.fields:
            field = getattr(self, field_name)
            check = field.validate()
            valid &= check["valid"]
            message += check["message"]
            excel_data[field_name] = check

            statuses.append(check["status"])

        if set(statuses) == {VALUE_UNCHANGED}:
            status = VALUE_UNCHANGED
        elif VALUE_CHANGED in statuses:
            status = VALUE_CHANGED
        else:
            status = VALUE_NEW

        excel_data["row_data"] = row_data
        data = {
            "valid": valid,
            "message": message,
            "count": len(self.rows),
            "excel_data": excel_data,
            "status": status,
        }
        return data

    def prepare_sheet(self, sheet):
        """To perform function prepare_sheet."""
        row = self.Meta.row_class
        for field_name in row.Meta.fields:
            field = getattr(row, field_name)
            if field.hidden and field_name not in self.visible_columns:
                sheet.column_dimensions[field.column].hidden = True
        return sheet

    def prepare_excel(self, supply_chain=None):
        """To perform function prepare_excel."""
        FARM_OPERATION_CHOICES = [
            o.name
            for o in Operation.objects.filter(
                node_type=constants.NODE_TYPE_FARM
            )
        ]
        # Filter connection type with supply chain and node type for
        # farmer bulk upload.
        if supply_chain:
            FARM_OPERATION_CHOICES = [
                o.name
                for o in Operation.objects.filter(
                    node_type=constants.NODE_TYPE_FARM,
                    supply_chains=supply_chain,
                )
            ]
        OPERATION_CHOICES = ",".join(FARM_OPERATION_CHOICES)
        OPERATION_CHOICES = '"' + OPERATION_CHOICES + '"'

        workbook = load_workbook(self.file_path)
        try:
            sheet = workbook[self.Meta.data_sheet]
        except KeyError:
            raise BadRequest("Invalid Excel file")

        sheet = self.prepare_sheet(sheet)
        for field_name in self.Meta.fields:
            field = getattr(self, field_name)
            if field.write:
                sheet[field.cell] = self.data[field_name]
        row_name = self.Meta.row_class.Meta.name
        if row_name in self.data:
            row = self.Meta.first_row
            for item in self.data[row_name]:
                row_object = self.Meta.row_class(
                    initial_data=item,
                    index=row,
                    visible_columns=self.visible_columns,
                )
                sheet = row_object.set_value(sheet)
                row += 1
        sheet.protection.sheet = True

        # A bad bad bad way to get this done. But tried several things, but
        # could not find a solution.
        for i in range(self.Meta.first_row, 501):
            formula1 = "=INDIRECT(%s!$%s$%d)" % ("Data", "EZ", i)
            val_subd_indirect = DataValidation(
                type="list",
                formula1=formula1,
                allowBlank=True,
                showErrorMessage=False,
                errorTitle="Invalid Sub Division",
                error="Please select a value from the dropdown.",
            )
            sheet.add_data_validation(val_subd_indirect)
            val_subd_indirect.add(sheet["%s%d" % ("J", i)])

            # if model_name name exists, the model_name should be
            # farmer and set connection type cell as dropdown with
            # operation choices of farmer node.
            if self.model_name:
                data_val = DataValidation(
                    type="list",
                    formula1=OPERATION_CHOICES,
                    allowBlank=True,
                    showErrorMessage=False,
                    errorTitle="Invalid Farmer type",
                    error="Invalid type. Select one from dropdown.",
                )
                sheet.add_data_validation(data_val)
                # set operation choices to cell E.
                data_val.add(sheet["%s%d" % ("E", i)])

        self.workbook = workbook

    def get_excel(self):
        """To perform function get_excel."""
        if self.file_path and not self.workbook:
            raise AssertionError(
                "You must call `.prepare_excel()` before you can get the"
                " excel."
            )
        data = save_virtual_workbook(self.workbook)
        return data
