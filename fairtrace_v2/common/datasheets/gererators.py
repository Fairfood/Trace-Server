import datetime
import os
import shutil
from operator import methodcaller
from typing import List

import pytz
from django.conf import settings
from django.core.files import File
from django.db import models
from django.utils import timezone
from openpyxl import load_workbook

DATASHEET_ROOT = "/common/datasheets/"

# Index for fetching value from cell tuples.
ROW_INDEX = 0
COL_INDEX = 1


class GeneratorMethod:
    """A 'GeneratorMethod' is used identify the field as a method field.

    * In 'Meta.fields' defines field as {'field_name': GeneratorMethod('foo')}

        def foo(self, instance):
            ...
            return ...
    """

    def __init__(self, name):
        """To perform function __init__."""
        self.name = name

    def __repr__(self):
        """To perform function __repr__."""
        return f"{self.name}"


class DataGenerator(object):
    """A 'DataGenerator' is a serializer mainly used for generating data for
    the excel/csv files. To instantiate a 'DataGenerator':

    * Inherit 'DataGenerator'
    * Add 'model' and 'fields_map' to class 'Meta'
    * We can access nested relation by binding related field with '__'
    (double underscore)
    * We can bind nested reverse related fields also, output will be a
    comma-separated string
    * We can use model properties as field and also bind the nested fields
    if it returns a model instance or a queryset.
    * If field is a method then use 'GeneratorMethod()' to get the field value.

    .data(queryset) -> This will generate a data 'dict' according to the
    fields mentioned in 'Meta'.

    .clean(field, value, data) -> This will clean individual fields value
    and helps to compare with other resolved fields.
    """

    meta = None
    fields = []
    new_field_names = []

    def __init__(self):
        """To perform function __init__."""
        self._set_meta_data()
        self._set_clean_fields()

    def _set_meta_data(self):
        """Check and validate Meta class."""
        assert hasattr(self, "Meta"), "No Meta class found."
        # Meta is available, we reference the meta instance to the self.
        self.meta = getattr(self, "Meta")
        # Model and fields are mandatory for 'Meta'
        assert hasattr(self.meta, "model"), "Meta has no model"
        assert hasattr(self.meta, "field_map"), "Meta has no field_map"

    def _set_clean_fields(self):
        """Clean and set fields to the generator."""
        self.fields = list(self.meta.field_map.values())
        self.new_field_names = list(self.meta.field_map.keys())

    def _model_check(self, model):
        """Check meta model is valid or not."""
        assert model == self.meta.model, "Meta model mismatch."

    def data(self, qs: models.QuerySet) -> List[dict]:
        """convert qs to dictionary."""
        self._model_check(qs.model)
        data_list = []
        for item in qs.iterator():
            data_dict = {}
            for field in self.fields:
                value = self.get_value(field, item)
                field = self.new_field_names[self.fields.index(field)]
                self.clean(field, value, data_dict)
            data_list.append(data_dict)
        return data_list

    # noinspection PyMethodMayBeStatic
    def clean(self, field, value, data):
        """Override this function to modify individual fields values and
        compare with already resolved fields in the data.

        * IMPORTANT: call 'super()' at the end, otherwise fields will not
        updated.
        """
        data[field] = value

    def get_model_value(self, field, item) -> object:
        """Get model related directly constructed field values."""
        if not item:
            return item
        field_seg = field.split("__")
        checks = (
            item.__class__.__name__ != "RelatedManager",
            not isinstance(item, models.QuerySet),
        )
        if all(checks):
            field = field_seg.pop(0)
            value = getattr(item, field)

            if not field_seg:
                # Skipping further resolving if value is None.
                # Retuning value only if there is no more segments to resolve.
                return value if value else None
            return self.get_model_value("__".join(field_seg), value)

        # Item maybe a RelatedManager or a QuerySet.
        values = []
        for obj in item.all():
            values.append(str(self.get_model_value(field, obj)))
        return ", ".join(values)

    def get_value(self, field, item):
        """Returns value by resolving the filed."""
        if isinstance(field, GeneratorMethod):
            return methodcaller(str(field), item)(self)
        try:
            value = self.get_model_value(field, item)
            # Handle QuerySet objects
            if isinstance(value, models.QuerySet):
                return ", ".join([str(v) for v in value])
        except AttributeError as err:
            raise AttributeError(f"{err} - key: '{field}'")
        return value


class TemplateDataGenerator(DataGenerator):
    """A 'TemplateDataGenerator' is inherited from 'DataGenerator', creates a
    data_sheet from a queryset with template.

    * template_name -> Reference template file name.
    * Data entering cell locations (x, y):
        - start_cell -> rows start cell.
        - created_by_cell -> created_by cell.
        - created_on_cell -> created_on cell.
        - company_cell -> company name cell.
    * You can override path while initializing.

    .to_sheet(queryset) -> create data sheet
    .save_to_instance() -> append data sheet to export entry.
    .close() -> remove temporary file from storage.
    """

    def __init__(self, instance=None, path=None):
        """To perform function __init__."""
        super(TemplateDataGenerator, self).__init__()
        self.copy_template()
        self.instance = instance
        if path:
            self.path = settings.BASE_DIR + path

    template_name = "default.xlsx"

    start_cell = (4, 1)
    created_by_cell = (1, 1)
    created_on_cell = (2, 1)
    company_cell = (3, 1)

    path = settings.BASE_DIR + DATASHEET_ROOT
    full_path = None

    def copy_template(self):
        """Copy and create a temporary file from template."""
        assert self.template_name is not None, "template_name cannot be empty."
        timestamp = str(timezone.now().timestamp()).split(".")[0]
        new_template_name = f"{timestamp}{self.template_name}"
        self.full_path = self.path + "tmp/" + new_template_name
        shutil.copy(
            self.path + "templates/" + self.template_name, self.full_path
        )

    def to_sheet(self, queryset):
        """This function get the data and generate the file in a temporary
        storage."""
        data = self.data(queryset)
        wb = load_workbook(self.full_path)

        sheet = wb.active
        self.update_sheet_meta(sheet)
        for row_idx, row in enumerate(data, self.start_cell[ROW_INDEX]):
            for column_idx, column in enumerate(
                row.values(), self.start_cell[COL_INDEX]
            ):
                if isinstance(column, datetime.datetime):
                    if settings.USE_TZ and timezone.is_aware(column):
                        # Convert aware datetime to the default time zone
                        # before casting them to dates (#17742).
                        default_timezone = timezone.get_default_timezone()
                        column = timezone.make_naive(column, default_timezone)

                sheet.cell(row=row_idx, column=column_idx).value = column
        wb.save(self.full_path)

    def save_to_instance(self, file_name):
        """This will append the file with the Export entry."""
        if not self.instance:
            raise Exception("No Export instance associated.")
        with open(self.full_path, "rb") as file:
            self.instance.file.save(file_name, File(file))

    def close(self):
        """This function close the operation and remove temporary file from
        storage."""
        if os.path.exists(self.full_path):
            os.remove(self.full_path)
        else:
            raise Exception("The file does not exist")

    def update_sheet_meta(self, sheet):
        """Individual cells updating for extra cases."""
        if not self.instance:
            return

        if self.created_by_cell:
            sheet.cell(
                row=self.created_by_cell[ROW_INDEX],
                column=self.created_by_cell[COL_INDEX],
            ).value = self.instance.creator.name
        if self.created_on_cell:
            sheet.cell(
                row=self.created_on_cell[ROW_INDEX],
                column=self.created_on_cell[COL_INDEX],
            ).value = self.instance.created_on.date()
        if self.company_cell:
            sheet.cell(
                row=self.company_cell[ROW_INDEX],
                column=self.company_cell[COL_INDEX],
            ).value = (
                self.instance.node.full_name if self.instance.node else ""
            )
