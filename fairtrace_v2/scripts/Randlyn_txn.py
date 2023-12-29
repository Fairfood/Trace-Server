from dateutil import parser
from openpyxl import load_workbook

src_book = load_workbook(filename="source.xlsx")
dst_book = load_workbook(filename="capuca.xlsx")
UNIT = "KG"
CURRENCY = "HNL"

SOURCE_DATA_ROW = 2
SOURCE_ID_COLUMN = 1

SOURCE_TXNDATE_COLUMN = 2
SOURCE_PPU_COLUMN = 5
SOURCE_QT_COLUMN = 6
SOURCE_INVOICE_COLUMN = 7
SOURCE_SHEET_NAME = "Sheet1"
IGNORE_IDS = []

DST_DATA_ROW = 8
DST_ID_COLUMN = 6
DST_TXNDATE_COLUMN = 23
DST_UNIT_COLUMN = 24
DST_CURRENCY_COLUMN = 25
DST_PPU_COLUMN = 26
DST_QT_COLUMN = 27
DST_INVOICE_COLUMN = 28


def copy_row(src_row, dst_row, sheet):
    """To perform function copy_row."""
    row = sheet[src_row]
    if dst_row == 545:
        print("******************************", row[1].value)
    for cell in row:
        sheet.cell(row=dst_row, column=row.index(cell) + 1).value = cell.value


def remove_empty_row(sheet):
    """To perform function remove_empty_row."""
    for row in sheet.iter_rows():
        if row[0].row < DST_DATA_ROW:
            continue
        print("date: ", row[DST_UNIT_COLUMN - 1].value)
        if not row[DST_UNIT_COLUMN - 1].value:
            print(
                "deleting :",
                row[2].value,
                row[3].value,
                row[DST_TXNDATE_COLUMN - 1].value,
            )
            sheet.delete_rows(int(row[0].row), 1)


dst_end_row = 0
source = src_book[SOURCE_SHEET_NAME]
dst = dst_book["Datasheet"]
ids = []
mapping = {}
for row in source.iter_rows():
    if row[0].row < SOURCE_DATA_ROW:
        continue
    ids.append((row[SOURCE_ID_COLUMN - 1].value))

for row in dst.iter_rows():
    if row[0].row < DST_DATA_ROW:
        print(row[1].value, row[0].row)
        continue
    if not (row[DST_ID_COLUMN - 1].value):
        dst_end_row = int(row[0].row)
        break
    if (row[DST_ID_COLUMN - 1].value) in ids:
        mapping[(row[DST_ID_COLUMN - 1].value)] = int(row[0].row)

missing = []
for row in source.iter_rows():
    if row[0].row < SOURCE_DATA_ROW:
        continue
    id = row[SOURCE_ID_COLUMN - 1].value
    if id in IGNORE_IDS:
        continue

    if not (id in mapping.keys()):
        missing.append(id)
        continue
    target_row = mapping[id]
    if dst.cell(row=target_row, column=DST_TXNDATE_COLUMN).value:
        # print('override id: ', id, ' to row ', dst_end_row)
        copy_row(target_row, dst_end_row, dst)
        target_row = dst_end_row
        dst_end_row += 1

    dst.cell(row=target_row, column=DST_TXNDATE_COLUMN).value = parser.parse(
        str(row[SOURCE_TXNDATE_COLUMN - 1].value)
    ).strftime("%d-%m-%Y")
    dst.cell(row=target_row, column=DST_UNIT_COLUMN).value = UNIT
    dst.cell(row=target_row, column=DST_CURRENCY_COLUMN).value = CURRENCY
    dst.cell(row=target_row, column=DST_PPU_COLUMN).value = row[
        SOURCE_PPU_COLUMN - 1
    ].value
    dst.cell(row=target_row, column=DST_QT_COLUMN).value = row[
        SOURCE_QT_COLUMN - 1
    ].value
    dst.cell(row=target_row, column=DST_INVOICE_COLUMN).value = row[
        SOURCE_INVOICE_COLUMN - 1
    ].value

dst_book.save("Output.xlsx")

print("dst_end_row: ", dst_end_row)
print("missing:", missing)
oc = 0
for mis in list(set(missing + IGNORE_IDS)):
    oc += ids.count(mis)
print("not added txns :", oc)
print("total txns :", len(ids))

print("missings ", (missing + IGNORE_IDS))
for id in ids:
    if id in list(set(missing + IGNORE_IDS)):
        print("txn issue in row", (ids.index(id) + 2 + SOURCE_DATA_ROW))
